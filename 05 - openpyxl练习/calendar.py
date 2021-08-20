import calendar
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.drawing.image import Image


# 指定一周的第一天 0是星期一 6是星期天
from openpyxl.utils import get_column_letter

calendar.setfirstweekday(firstweekday=6)
year = 2021

# 创建excel
wb = openpyxl.Workbook()

# 循环月份
for i in range(1, 13):
    # 添加工作表 每个月份对应一个工作表
    sheet = wb.create_sheet(index=0, title=str(i) + "月")
    # 每月中的第一行 一行表示一周
    for j in range(len(calendar.monthcalendar(year, i))):
        # 每一天
        for k in range(len(calendar.monthcalendar(year, i)[j])):
            # 具体的日期
            value = calendar.monthcalendar(year, i)[j][k]
            if value == 0:
                # 将0值变为空值 没有日期的单元格天空值
                value = ''
                sheet.cell(row=j + 9, column=k + 1).value = value
            else:
                # 将日期数据添加到具体的单元格中
                sheet.cell(row=j + 9, column=k + 1).value = value
                # 设置字体
                sheet.cell(row=j + 9, column=k + 1).font = Font("微软雅黑", size=11)
    # 单元格文字设置 右对齐 垂直居中
    align = Alignment(horizontal="right", vertical="center")
    # 对单元格填充色属性设置
    fill = PatternFill("solid", fgColor="99CCCC")
    # 对单元格进行颜色填充`   1qq7uuu
    for k1 in range(1, 50):
        for k2 in range(1, 50):
            sheet.cell(row=k1, column=k2).fill = fill

    # 图片
    sheet.merge_cells("I1:P20")
    img = Image("./1.jpg")
    # 设置图片大小
    newSize = (200, 200)
    img.width, img.height = newSize
    # 顶部加一些距离
    sheet.add_image(img, "I2")

    # 星期日开头
    days = ["星期日", "星期一", "星期二", "星期三", "星期四", "星期五", "星期六"]
    num = 0
    # 添加星期几相关信息
    for k3 in range(1, 8):
        sheet.cell(row=8, column=k3).value = days[num]
        # 设置样式
        sheet.cell(row=8, column=k3).alignment = align
        sheet.cell(row=8, column=k3).font = Font("微软雅黑", size=11)
        # 设置列宽12
        c_char = get_column_letter(k3)
        sheet.column_dimensions[c_char].width = 12
        num += 1

    # 将日历所在单元格的行高都修改为30
    for k4 in range(8, 14):
        sheet.row_dimensions[k4].height = 30

    # 添加年份及月份
    sheet.cell(row=3, column=1).value = f"{year}年"
    sheet.cell(row=4, column=1).value = str(i) + "月"
    # 设置年份及月份文本样式
    sheet.cell(row=3, column=1).font = Font("微软雅黑", size=16, bold=True, color="FF7887")
    sheet.cell(row=4, column=1).font = Font("微软雅黑", size=16, bold=True, color="FF7887")
    # 设置年份及月份文本对齐方式
    sheet.cell(row=3, column=1).alignment = align
    sheet.cell(row=4, column=1).alignment = align

wb.save("./calendar.xlsx")
