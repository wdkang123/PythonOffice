import openpyxl
from openpyxl.styles import Font, colors, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment


# -------- 创建工作簿 ------------
# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
print(f"ws: title {ws.title}")

# 新建一个工作簿
ws2 = wb.create_sheet("new_sheet", 1)
# 修改title
ws2.title = "update_sheet"
# 添加内容
ws2.cell(row=5, column=1).value = "设置了一个值"

# 保存excel
wb.save("./openpyxl_excel.xlsx")

# --------- 修改工作簿 -----------
# 修改工作簿
wb = openpyxl.load_workbook("./openpyxl_excel.xlsx")
ws = wb.active

# 修改A2-C3区域的值
for row in ws['A2': 'C3']:
    for cell in row:
        cell.value = "修改了的值"

wb.save("./openpyxl_excel.xlsx")


# ---------- 修改工作簿单元格式 -----------
wb = openpyxl.Workbook()
ws = wb.active

rows = [
    ["id", "name", "age"],
    [1, "张三", 13],
    [2, "李四", 15],
    [3, "王五", 18]
]

for row in rows:
    # 添加到ws中
    ws.append(row)


# Font 字体设置微软雅黑 字体大小25 斜体 红色
font = Font(name="微软雅黑", size=25, italic=True, color=colors.RED, bold=True)
# 设置对应单元格的字体样式
ws['A1'].font = font

# PatternFill 填充单元格背景色填充绿色
fill = PatternFill(fill_type="solid", start_color=colors.GREEN)
ws['B1'].fill = fill

# Border Side 设置单元格边框样式
border = Border(
    left=Side(border_style="double", color="FFBB00"),
    right=Side(border_style="double", color="FFBB01"),
    top=Side(border_style="double", color="FFBB02"),
    bottom=Side(border_style="double", color="FFBB03")
)
ws['C1'].border = border

# Alignment 设置对其格式
align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 修改单元格高度 和 长度
ws.row_dimensions[3].height = 40
ws.column_dimensions["A"].width = 30

# 合并一行中的几个单元格
ws.merge_cells("A7:C7")
# 合并一个矩形区域中的单元格
ws.merge_cells("A9:C13")
ws["A9"] = "合并单元格"
ws["A9"].comment = Comment(text="这是一个批注", author="二两")

wb.save("./openpyxl_style_excel.xlsx")
